"""Convert PropStream SOLD exports into Miami Comp Engine SaleRow batches.

This importer is intentionally conservative:
- only SOLD rows are included by default
- sale price prefers Last Sale Amount over MLS Amount
- verified remains false because PropStream is a secondary export, not the
  county record itself
- generated JSON is meant for review before adding --push
"""
import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SALE_BATCH_LIMIT = 500
SFR_TYPES = {"single family residential", "residential (general) (single)"}
WATER_STREET_HINTS = (
    " BAY ",
    "BAY DR",
    "BROADVIEW",
    "KEYSTONE",
    "HIBISCUS",
    "PALM",
    "STAR ISLAND",
    "SUNSET ISLAND",
    "DI LIDO",
    "RIVO ALTO",
    "SAN MARINO",
    "LA GORCE CIR",
    "ALLISON",
    "N SHORE DR",
)


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = clean_text(value)
    return int(re.sub(r"[^\d]", "", text)) if re.search(r"\d", text) else None


def clean_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        text = re.sub(r"[^0-9.]", "", str(value))
        return float(text) if text else None


def date_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def date_key(value):
    iso = date_iso(value)
    return iso or ""


def folio(value):
    text = clean_text(value)
    return re.sub(r"[^0-9]", "", text)


def infer_micro_market(row):
    city = clean_text(row.get("City")).upper()
    zip_code = clean_text(row.get("Zip"))
    address = clean_text(row.get("Address")).upper()
    if city == "BAY HARBOR ISLANDS":
        return "bay_harbor_islands"
    if city in {"BAL HARBOUR", "BAL HARBOR"}:
        return "bal_harbour"
    if city == "SURFSIDE":
        return "surfside"
    if city == "NORTH MIAMI" or zip_code == "33181" or "KEYSTONE" in address:
        return "keystone"
    if city == "NORTH MIAMI BEACH":
        return "north_miami_beach"
    if city == "MIAMI BEACH":
        return "miami_beach"
    return city.lower().replace(" ", "_") or "unknown"


def infer_water_type(row):
    text = f" {clean_text(row.get('Address')).upper()} {clean_text(row.get('City')).upper()} "
    if any(hint in text for hint in WATER_STREET_HINTS):
        return "Verify"
    return "None"


def condition_class(row):
    condition = clean_text(row.get("Total Condition")).lower()
    if condition in {"luxury", "excellent", "good"}:
        return "renovated"
    if condition in {"average", "fair"}:
        return "original"
    if condition in {"poor", "very poor"}:
        return "teardown"
    return "unknown"


def row_dicts(path):
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for cells in ws.iter_rows(min_row=2):
        values = [cell.value for cell in cells]
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in values):
            yield row


def sale_price_date_and_source(row, min_sale_date):
    last_sale_date = date_key(row.get("Last Sale Recording Date"))
    last_sale_amount = clean_int(row.get("Last Sale Amount"))
    if last_sale_date >= min_sale_date and last_sale_amount:
        return last_sale_amount, last_sale_date, "propstream_last_sale"

    mls_date = date_key(row.get("MLS Date"))
    mls_amount = clean_int(row.get("MLS Amount"))
    if mls_date >= min_sale_date and mls_amount:
        return mls_amount, mls_date, "propstream_mls_sold"

    return None, None, "missing_recent_sale"


def normalize_sale(row, source_name, args):
    status = clean_text(row.get("MLS Status")).upper()
    if args.sold_only and status != "SOLD":
        return None, "not sold"

    prop_type = clean_text(row.get("Property Type"))
    if args.sfr_only and prop_type.lower() not in SFR_TYPES:
        return None, f"not SFR: {prop_type}"

    price, sale_date, price_source = sale_price_date_and_source(row, args.min_sale_date)
    if not price:
        return None, price_source
    if price < args.min_price:
        return None, f"below min price: {price}"

    apn = clean_text(row.get("APN"))
    clean_folio = folio(apn)
    address = clean_text(row.get("Address"))
    unit = clean_text(row.get("Unit #"))
    city = clean_text(row.get("City"))
    zip_code = clean_text(row.get("Zip"))
    row_id_base = clean_folio or re.sub(r"[^a-z0-9]+", "-", address.lower()).strip("-")
    sale_id = f"propstream:{row_id_base}:{sale_date.replace('-', '')}:{price}"

    return {
        "saleId": sale_id,
        "address": address,
        "unit": unit,
        "city": city,
        "zip": zip_code,
        "salePrice": price,
        "saleDate": sale_date,
        "livingSf": clean_int(row.get("Building Sqft")),
        "lotSf": clean_int(row.get("Lot Size Sqft")),
        "yearBuilt": clean_int(row.get("Effective Year Built")),
        "beds": clean_int(row.get("Bedrooms")),
        "baths": clean_float(row.get("Total Bathrooms")),
        "propertyClass": "sfr",
        "waterType": args.water_type or infer_water_type(row),
        "conditionClass": args.condition_class or condition_class(row),
        "subdivision": "",
        "microMarket": args.micro_market or infer_micro_market(row),
        "verified": False,
        "source": f"{args.source}:{price_source}:{source_name}",
    }, ""


def write_json(path, payload):
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2, default=str))


def write_batches(rows, output_path):
    write_json(output_path, rows)
    batch_paths = []
    if len(rows) > SALE_BATCH_LIMIT:
        stem = Path(output_path).with_suffix("")
        for idx in range(0, len(rows), SALE_BATCH_LIMIT):
            batch_path = Path(f"{stem}_batch_{idx // SALE_BATCH_LIMIT + 1:03d}.json")
            write_json(batch_path, rows[idx:idx + SALE_BATCH_LIMIT])
            batch_paths.append(batch_path)
    return batch_paths


def push_batches(rows):
    from engine_client import engine_request

    results = []
    for idx in range(0, len(rows), SALE_BATCH_LIMIT):
        batch = rows[idx:idx + SALE_BATCH_LIMIT]
        print(f"Pushing batch {idx // SALE_BATCH_LIMIT + 1}: {len(batch)} rows", flush=True)
        results.append(engine_request("/api/import-sales", {"rows": batch}))
    return results


def build_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", nargs="+", help="PropStream SOLD export .xlsx file(s).")
    parser.add_argument("--out", default="propstream_sales_import.json")
    parser.add_argument("--push", action="store_true")
    parser.add_argument("--sold-only", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--sfr-only", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--min-price", type=int, default=100000)
    parser.add_argument("--min-sale-date", default="2024-01-01", help="Oldest sale date to import.")
    parser.add_argument("--source", default="propstream_recent_sold")
    parser.add_argument("--water-type", choices=("Canal", "Open Bay", "Canal/Bay", "None", "Verify"))
    parser.add_argument("--condition-class", choices=("renovated", "original", "teardown", "unknown"))
    parser.add_argument("--micro-market")
    return parser


def main():
    args = build_parser().parse_args()
    rows = []
    skipped = []

    for file_name in args.xlsx:
        path = Path(file_name)
        for idx, row in enumerate(row_dicts(path), start=2):
            normalized, reason = normalize_sale(row, path.stem, args)
            if normalized:
                rows.append(normalized)
            else:
                skipped.append({"file": str(path), "row": idx, "reason": reason})

    deduped = {row["saleId"]: row for row in rows}
    rows = list(deduped.values())
    batch_paths = write_batches(rows, args.out)
    if skipped:
        write_json(Path(args.out).with_suffix(".skipped.json"), skipped)

    print(f"Rows ready: {len(rows)}")
    print(f"Skipped rows: {len(skipped)}")
    print(f"Wrote: {args.out}")
    for batch_path in batch_paths:
        print(f"Wrote batch: {batch_path}")

    if args.push:
        print(json.dumps(push_batches(rows), indent=2))
    else:
        print("Dry run only. Add --push after reviewing the output JSON.")


if __name__ == "__main__":
    main()
